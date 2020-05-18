import openpyxl

new_wb = openpyxl.load_workbook('data.xlsx')
origin_wb = openpyxl.load_workbook('testset.xlsx')

new_sheet = new_wb.active
origin_sheet = origin_wb.active

#for j in range (0, 25):
#	for i in range (0, 180):
#		new_sheet.cell(row=2+i, column=1+j).value = origin_sheet.cell(row=3+j+26*i, column=4).value

year = 2003

for k in range(0, 180):
	if k % 12 == 0:
		year += 1
	month = k%12+1
	new_sheet.cell(row=2+k, column=26).value = str(year) + '-' + str(month)

new_wb.save('data.xlsx')
